import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook
from PIL import Image, ImageTk
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk  
import threading  
import os
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

driver = None
processando = False


def carregar_planilha(caminho_planilha):
    df = pd.read_excel(caminho_planilha)
    if 'Status' not in df.columns:
        df['Status'] = ''  
    return df


def iniciar_driver():
    global driver
    if driver is None:
        try:
            options = webdriver.ChromeOptions()
            options.add_argument("--headless")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            
            user_data_dir = os.path.expanduser("~") + "/AppData/Local/Google/Chrome/User Data"
            options.add_argument(f"user-data-dir={user_data_dir}")
            options.add_experimental_option("debuggerAddress", "localhost:9222")

            driver = webdriver.Chrome(options=options)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
            print("Driver iniciado com sucesso!")
        except Exception as e:
            print(f"Erro ao iniciar o driver: {e}")
            driver = None
    else:
        try:
            driver.execute_script("return document.readyState")
        except Exception as e:
            print(f"Driver desconectado, tentando reconectar: {e}")
            driver = None
            iniciar_driver()

    return driver


def verificar_e_mudar_url(driver, url_desejada):
    url_atual = driver.current_url
    if url_atual != url_desejada:
        print(f"Navegando para a URL desejada: {url_desejada}")
        driver.get(url_desejada)
    else:
        print(f"Já está na URL correta: {url_atual}")

def verificar_cliente(driver, cnpj, contador):
    try:
        print(f"{contador} Preenchendo o CNPJ...")
        cnpj_input = WebDriverWait(driver, 1).until(
            EC.presence_of_element_located((By.XPATH, "//input[contains(@id, '708:0')]"))
        )
        cnpj_input.clear()
        cnpj_input.send_keys(cnpj)

        submit_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'slds-button') and contains(@class, 'uiButton--brand') and span[text()='Confirmar']]"))
        )
        submit_button.click()

        driver.execute_script("document.querySelector('.loadingCon.global.siteforceLoadingBalls').style.display = 'none';")
        WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.XPATH, "//div[contains(@class, 'loadingCon')]")))

        try:
            erro_telefone = WebDriverWait(driver, 0.5).until(
                EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), 'O formato correto para o telefone é DDI + DDD + telefone')]"))
            )
            if erro_telefone:
                return "LIVRE"
        except TimeoutException:
            pass

        try:
            erro_cadastro = WebDriverWait(driver, 0.5).until(
                EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), 'Já existe um lead cadastrado com o CNPJ informado.')]"))
            )
            if erro_cadastro:
                return "CARIMBADO"
        except TimeoutException:
            pass

        try:
            erro_cliente = WebDriverWait(driver, 0.5).until(
                EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), 'Já existe um lead e um cliente cadastrado com o CNPJ informado.')]"))
            )
            if erro_cliente:
                return "JÁ É CLIENTE"
        except TimeoutException:
            pass

        return "CARIMBADO"
    except NoSuchElementException as e:
        print(f"Erro ao localizar elemento: {e}")
        return "ERRO: Elemento não encontrado"
    except TimeoutException as e:
        print(f"Erro de tempo de espera: {e}")
        return "ERRO: Tempo esgotado"

def criar_nova_planilha(df, caminho_planilha):
    workbook = load_workbook(caminho_planilha)
    sheet = workbook.active
    colunas = list(df.columns)
    coluna_status = len(colunas)

    for index, status in enumerate(df['Status'], start=2):
        sheet.cell(row=index, column=coluna_status).value = status

    novo_caminho = caminho_planilha.replace('.xlsx', '_resultado.xlsx')
    workbook.save(novo_caminho)

def processar_verificacao(caminho_planilha, progress_bar, progress_label):
    global processando
    processando = True
    df = carregar_planilha(caminho_planilha)
    driver = iniciar_driver()

    if driver is None:
        messagebox.showerror("Erro", "Não foi possível iniciar o driver do navegador.")
        return

    total = len(df)
    contador = 1

    for index, row in df.iterrows():
        if not processando:
            break
        
        cnpj = row['CNPJ']
        status = verificar_cliente(driver, cnpj, contador)
        df.at[index, 'Status'] = status
        
        progress = (contador / total) * 100
        progress_bar['value'] = progress
        progress_label.config(text=f"{contador}/{total} ({int(progress)}%)")
        progress_bar.update_idletasks()

        contador += 1

    criar_nova_planilha(df, caminho_planilha)

    messagebox.showinfo("Concluído", "Processamento da planilha finalizado!")

    progress_bar['value'] = 0
    progress_label.config(text="0/0 (0%)")
    
def cancelar_processamento():
    global processando
    processando = False
    messagebox.showinfo("Cancelado", "Processamento foi cancelado!")

def selecionar_planilha(progress_bar, progress_label):
    caminho_planilha = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if caminho_planilha:
        threading.Thread(target=processar_verificacao, args=(caminho_planilha, progress_bar, progress_label)).start()

def criar_interface():
    root = tk.Tk()
    root.title("BauBot :)")
    root.geometry("400x250")
    root.resizable(False, False)
    
    icon_image = ImageTk.PhotoImage(Image.open("logo.ico"))  
    root.iconphoto(True, icon_image)

    fonte_padrao = ("Calibri", 12)

    label = tk.Label(root, text="Selecione a planilha para verificar:", font=fonte_padrao)
    label.pack(pady=20)

    progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
    progress_bar.pack(pady=10)

    progress_label = tk.Label(root, text="0/0 (0%)", font=fonte_padrao)
    progress_label.pack()

    botao_selecionar = tk.Button(root, text="Selecionar Planilha", command=lambda: selecionar_planilha(progress_bar, progress_label), font=fonte_padrao)
    botao_selecionar.pack(pady=10)
    
    botao_cancelar = tk.Button(root, text="Cancelar Processamento", command=cancelar_processamento, font=fonte_padrao)
    botao_cancelar.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    criar_interface()