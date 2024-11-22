import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk  
import threading  
import os

drivers = []  # Lista para armazenar os drivers
processando = False

def carregar_planilha(caminho_planilha):
    df = pd.read_excel(caminho_planilha)
    if 'Status' not in df.columns:
        df['Status'] = ''  
    return df

def iniciar_drivers():
    global drivers
    for i in range(3):  # Iniciar 3 drivers
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")  # Mantenha como está se não precisar ver o navegador
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        user_data_dir = os.path.expanduser("~") + "/Users/ALCANCE/Desktop/Login Chrome"
        options.add_argument(f"user-data-dir={user_data_dir}")
        options.add_experimental_option("debuggerAddress", "localhost:9222")
        
        driver = webdriver.Chrome(options=options)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
        drivers.append(driver)

def dividir_nome_completo(nome_completo):
    partes = nome_completo.split()
    nome = partes[0]  # Primeiro nome
    sobrenome = " ".join(partes[1:]) if len(partes) > 1 else ""  # O resto é o sobrenome
    return nome, sobrenome

def verificar_cliente(driver, nome, sobrenome, email, telefone, cnpj, contador):
    try:
        print(f"{contador} Preenchendo os dados...")

        # Preencher Nome
        nome_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[contains(@id, '610:0')]")))
        nome_input.clear()
        nome_input.send_keys(nome)

        # Preencher Sobrenome
        sobrenome_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[contains(@id, '620:0')]")))
        sobrenome_input.clear()
        sobrenome_input.send_keys(sobrenome)

        # Preencher Email
        email_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[contains(@id, '655:0')]")))
        email_input.clear()
        email_input.send_keys(email)

        # Preencher Telefone
        telefone_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[contains(@id, '671:0')]")))
        telefone_input.clear()
        telefone_input.send_keys(telefone)

        # Preencher CNPJ
        cnpj_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[contains(@id, '708:0')]")))
        cnpj_input.clear()
        cnpj_input.send_keys(cnpj)

        # Submeter o formulário
        submit_button = driver.find_element(By.XPATH, "//button[contains(@class, 'slds-button') and contains(@class, 'uiButton--brand') and span[text()='Confirmar']]")
        submit_button.click()

        # Espera até que o botão de confirmação esteja novamente clicável
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'slds-button') and contains(@class, 'uiButton--brand') and span[text()='Confirmar']]"))
        )

        # Verificar se o telefone está correto
        try:
            erro_telefone = WebDriverWait(driver, 6).until(
                EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), 'O formato correto para o telefone é DDI + DDD + telefone')]"))
            )
            if erro_telefone:
                return "LIVRE"
        except:
            return "CARIMBADO"

        return "CARIMBADO"

    except NoSuchElementException as e:
        print(f"Erro ao localizar elemento: {e}")
        return "ERRO: Elemento não encontrado"
    except TimeoutException as e:
        print(f"Erro de tempo de espera: {e}")
        return "ERRO: Tempo esgotado"

def criar_nova_planilha(df, caminho_planilha):
    workbook = load_workbook(caminho_planilha)
    sheet = workbook.active
    colunas = list(df.columns)
    coluna_status = len(colunas)

    for index, status in enumerate(df['Status'], start=2):
        sheet.cell(row=index, column=coluna_status).value = status

    novo_caminho = caminho_planilha.replace('.xlsx', '_resultado.xlsx')
    workbook.save(novo_caminho)

def processar_verificacao(caminho_planilha, progress_bar, progress_label):
    global processando
    processando = True
    df = carregar_planilha(caminho_planilha)
    iniciar_drivers()

    if not drivers:
        messagebox.showerror("Erro", "Não foi possível iniciar os drivers do navegador.")
        return

    total = len(df)
    contador = 0

    while contador < total and processando:
        for i in range(3):  # Usar até 3 drivers simultaneamente
            if contador < total:
                row = df.iloc[contador]  # Pegar a linha atual
                nome_completo = row['NOME']
                nome, sobrenome = dividir_nome_completo(nome_completo)
                email = row['EMAIL']
                telefone = row['TELEFONE']
                cnpj = row['CNPJ']

                # Usar o driver correspondente à aba
                driver = drivers[i]  # Alternar entre os 3 drivers
                status = verificar_cliente(driver, nome, sobrenome, email, telefone, cnpj, contador + 1)
                df.at[contador, 'Status'] = status

                progress = ((contador + 1) / total) * 100
                progress_bar['value'] = progress
                progress_label.config(text=f"{contador + 1}/{total} ({int(progress)}%)")
                progress_bar.update_idletasks()

                contador += 1  # Incrementar o contador

    criar_nova_planilha(df, caminho_planilha)

    messagebox.showinfo("Concluído", "Processamento da planilha finalizado!")

    progress_bar['value'] = 0
    progress_label.config(text="0/0 (0%)")

def cancelar_processamento():
    global processando
    processando = False
    messagebox.showinfo("Cancelado", "Processamento foi cancelado!")

def selecionar_planilha(progress_bar, progress_label):
    caminho_planilha = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if caminho_planilha:
        threading.Thread(target=processar_verificacao, args=(caminho_planilha, progress_bar, progress_label)).start()

def criar_interface():
    root = tk.Tk()
    root.title("BauBot :)")
    root.geometry("400x250")
    root.resizable(False, False)

    fonte_padrao = ("Calibri", 12)

    label = tk.Label(root, text="Selecione a planilha para verificar:", font=fonte_padrao)
    label.pack(pady=20)

    progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
    progress_bar.pack(pady=10)

    progress_label = tk.Label(root, text="0/0 (0%)", font=fonte_padrao)
    progress_label.pack()

    botao_selecionar = tk.Button(root, text="Selecionar Planilha", command=lambda: selecionar_planilha(progress_bar, progress_label), font=fonte_padrao)
    botao_selecionar.pack(pady=10)
    
    botao_cancelar = tk.Button(root, text="Cancelar Processamento", command=cancelar_processamento, font=fonte_padrao)
    botao_cancelar.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    criar_interface()